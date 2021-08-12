import sqlite3
import tkinter
from tkinter import *
from tkinter import messagebox
from tkinter import Entry
from tkinter import Listbox
from tkinter import Scrollbar
import os 
import cv2
import numpy as np
#import sys
from PIL import Image
import pickle
from tkinter import ttk
from db import Database
import openpyxl
from openpyxl.styles import Font,Alignment
import datetime


db=Database('store.db')
window=tkinter.Tk()
window.title("online attendance management system")

def recognize():
   
    
    #name
    name_text=tkinter.StringVar()
    namelabel=ttk.Label(window,text="Name")
    namelabel.grid(row=11,column=0)
    nameentry=ttk.Entry(window,width=26,textvariable=name_text)
    nameentry.grid(row=11,column=1)

    #regi
    regi_text=tkinter.StringVar()
    regilabel=ttk.Label(window,text="Regi")
    regilabel.grid(row=11,column=2)
    regilabel=ttk.Entry(window,width=26,textvariable=regi_text)
    regilabel.grid(row=11,column=3)

    #branch
    branch_text=tkinter.StringVar()
    branchlabel=ttk.Label(window,text="Branch")
    branchlabel.grid(row=11,column=4)
    branchentry=ttk.Entry(window,width=26,textvariable=branch_text)
    branchentry.grid(row=11,column=5)
    
    def action():
        messagebox.showinfo('warning',"your response has been recorded")
        db.insert(name_text.get(),regi_text.get(),branch_text.get())
        
        uid=regi_text.get()
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        if os.path.exists(BASE_DIR+'/'+'attendance.xlsx'):
            pass
        else:
            filepath = "/home/ubun/Desktop/stocksinfo/test101.xlsx"
            wb = openpyxl.Workbook()
            wb.save(filepath)

        Wkbpath = BASE_DIR+ "/"+"attendance.xlsx"
        wkb = openpyxl.load_workbook(Wkbpath)   
        
        sh=wkb['Sheet']
        x=3
        while (True):
            data=sh.cell(row=x,column=1).value
            if (data==None):
                sh.cell(row=x,column=1).value=str(uid)
                wkb.save(Wkbpath)
                break
            else:
                x=x+1    
                    

        attendance_list.delete(0,END)
        attendance_list.insert(END,(name_text.get(),regi_text.get(),branch_text.get()))
        populate_list()
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        face_classifier = cv2.CascadeClassifier(BASE_DIR+'/'+'haarcascade_frontalface_default.xml')


        def face_extractor(img):
            gray=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
            faces=face_classifier.detectMultiScale(gray,1.3,5)
                
            if faces is ():
                return None
            for(x,y,w,h) in faces:
                cropped_face=img[y:y+h,x:x+w]
            return cropped_face
        namefolder=regi_text.get()
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        if os.path.exists(BASE_DIR+'/'+'faces'):
            pass
        else :
            os.mkdir(BASE_DIR+'/'+'faces')
        os.chdir(BASE_DIR+'/'+'faces')

        os.mkdir(namefolder)


        cap=cv2.VideoCapture(0)
        count=0
        while True:
            

            ret,frame=cap.read()
            if face_extractor(frame) is not None:
                count=count+1
                face=cv2.resize(face_extractor(frame),(450,450))
                face=cv2.cvtColor(face,cv2.COLOR_BGR2GRAY)
                image_dir =os.path.join(BASE_DIR,'faces'+'/'+namefolder+'/'+str(count)+'.jpg') 
                #file_name_path='C:/Users/BEAST/Desktop/project/faces/faces'+str(count)+'.jpg'
                cv2.imwrite(image_dir,face)
                cv2.putText(face,str(count),(50,50),cv2.FONT_HERSHEY_COMPLEX,1,(0,255,0),2)
                cv2.imshow('face cropper',face)
            else:
                pass
            if cv2.waitKey(1)==13 or count==100:
                break
        cap.release()
        cv2.destroyAllWindows() 
        messagebox.showinfo('registraion',"Collecting samples completed")  

        
        
        
    btn=ttk.Button(window,text="submit",width=12,command=action)
    btn.grid(row=12,column=3)

    
 
    
    
   
    
def train():
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    image_dir =os.path.join(BASE_DIR,"faces") #file name where image is located

    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    face_cascade = cv2.CascadeClassifier(BASE_DIR+'/'+'haarcascade_frontalface_default.xml')

    
    recognizer=cv2.face.LBPHFaceRecognizer_create()

    current_id=0
    label_ids={}

    x_train=[]
    y_label=[]

    for root,dirs,files in os.walk(image_dir):
        for file in files:
            if file.endswith("png") or file.endswith("jpg"):
                path=os.path.join(root,file)
                label=os.path.basename(root).replace(" ","-").lower()
                if not label in label_ids:
                    label_ids[label]=current_id
                    current_id += 1

                id_=label_ids[label]
                pil_image = Image.open(path).convert("L")   #convert to grayscale
                size=(450,450)
                final_image=pil_image.resize(size,Image.ANTIALIAS)
                
                image_array=np.array(final_image,"uint8")
                faces = face_cascade.detectMultiScale(image_array, scaleFactor=1.5, minNeighbors=5)
            
                for (x,y,w,h) in faces:
                    roi=image_array[y:y+h,x:x+w]
                    x_train.append(roi)
                    y_label.append(id_)

                    
    with open (BASE_DIR+'/'+'faces'+'/'+"labels.pickle",'wb') as f:
        pickle.dump(label_ids,f)

    recognizer.train(x_train,np.array(y_label))
    recognizer.save(BASE_DIR+'/'+'faces'+'/'+"trainer.yml")
    messagebox.showinfo('warning',"image training completed")


def take_attendance():
   
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    face_cascade = cv2.CascadeClassifier(BASE_DIR+'/'+'haarcascade_frontalface_default.xml')

    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.read(BASE_DIR+'/'+'faces'+'/'+"trainer.yml")


    labels = {"person_name": 1}
    with open(BASE_DIR+'/'+'faces'+'/'+"labels.pickle", 'rb') as f:
        og_labels = pickle.load(f)
        labels = {v:k for k,v in og_labels.items()}

    cap = cv2.VideoCapture(0)

    while(True):
        # Capture frame-by-frame
        name=''
        ret, frame = cap.read()
        gray  = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.5, minNeighbors=5)
        for (x, y, w, h) in faces:
            #print(x,y,w,h)
            roi_gray = gray[y:y+h, x:x+w] #(ycord_start, ycord_end)
            roi_color = frame[y:y+h, x:x+w]
            id_, conf = recognizer.predict(roi_gray)
            if conf>=4 and conf <= 85:
                #print(5: #id_)
                #print(labels[id_])
                font = cv2.FONT_HERSHEY_SIMPLEX
                name = labels[id_]
                color = (255, 255, 255)
                stroke = 2
                cv2.putText(frame, name, (x,y), font, 1, color, stroke, cv2.LINE_AA)

            color = (255, 0, 0) #BGR 0-255 
            stroke = 2
            end_cord_x = x + w
            end_cord_y = y + h
            cv2.rectangle(frame, (x, y), (end_cord_x, end_cord_y), color, stroke)
            
        # Display the resulting frame
        cv2.imshow('frame',frame)
        Wkbpath = BASE_DIR+ "/"+"attendance.xlsx"
        wkb = openpyxl.load_workbook(Wkbpath)
        now=datetime.datetime.now()
       
        dt = datetime.date(year=now.year, month=now.month, day=now.day)  
        
        year=now.year
    
        
        sh=wkb[f"{dt:%B}" + "  " + str(year)]
        dt=f"{dt:%B %d, %Y}" 
        x=1
        for x in range (1,sh.max_column):
            if (sh.cell(row=1,column=x).value==dt):
                
                x=x+1
                break
        x_column=x-1
              
        for y in range (3,10):
            
            
            if(sh.cell(row=y,column=1).value==name):
                
                sh.cell(row=y,column=x_column).value='p'   
                wkb.save(Wkbpath)
            y=y+1    
                
                        
                    
                    
        if cv2.waitKey(20) & 0xFF == ord('q'):
            break

    



    # When everything done, release the capture
    cap.release()
    cv2.destroyAllWindows()

def generate_excel():
    
    #Define the workbook path and Load the workbook
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    Wkbpath = BASE_DIR+ "/"+"attendance.xlsx"
    wkb = openpyxl.load_workbook(Wkbpath)

    #Defined variables for Months and Year, and stored the values through input function
    #minmonth
    minmonth_text=tkinter.StringVar()
    minmonthlabel=ttk.Label(window,text="Minmonth")
    minmonthlabel.grid(row=11,column=0)
    minmonthentry=ttk.Entry(window,width=26,textvariable=minmonth_text)
    minmonthentry.grid(row=11,column=1)

    #maxmonth
    maxmonth_text=tkinter.StringVar()
    maxmonthlabel=ttk.Label(window,text="Maxmonth")
    maxmonthlabel.grid(row=11,column=2)
    maxmonthlabel=ttk.Entry(window,width=26,textvariable=maxmonth_text)
    maxmonthlabel.grid(row=11,column=3)

    #year
    year_text=tkinter.StringVar()
    yearlabel=ttk.Label(window,text="year")
    yearlabel.grid(row=11,column=4)
    yearentry=ttk.Entry(window,width=26,textvariable=year_text)
    yearentry.grid(row=11,column=5) 

    def action():
        MinMonth=int(minmonth_text.get())
        MaxMonth=int(maxmonth_text.get())
        Year=int(year_text.get())
        
        # Defined two lists -- (i) Day names in a week  (ii) Max number of days in a month
        DaysList = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        DaysCount = [31,28,31,30,31,30,31,31,30,31,30,31]

        # Defined variable for Input Sheet
        InputSH = wkb['Sheet']

        # Applied font properties
        fontstyle = Font(name = 'Century',size = 11, bold = True)

        try:
            if MinMonth>MaxMonth:
                raise ValueError()
        

            for sheet in  wkb.sheetnames:
                worksh =  wkb[sheet]
                if  worksh.title !=  'Sheet':
                    del  wkb[worksh.title]  

            for MonthNumb in range(MinMonth,MaxMonth+1):         
                SH = wkb.create_sheet(index = len(wkb.sheetnames)+1)
                SH['A1'].value = 'Date'; SH['A1'].font =  fontstyle
                SH['A2'].value = 'Day'; SH['A2'].font =  fontstyle
                        
                for r in range(InputSH.max_row):        
                    if r >=2:
                        SH.cell(row = r+1, column = 1).value = InputSH.cell(row = r+1, column = 1).value
                    for DayNumb in range(1,DaysCount[MonthNumb-1]+1):           
                            
                            dt = datetime.date(year=Year, month=MonthNumb, day=DayNumb)                    
                            if r == 0:
                                SH.cell(row = r+1, column = DayNumb+1).value = f"{dt:%B %d, %Y}"            
                                print(f"{dt:%B %d, %Y}")
                                SH.cell(row = r+1, column = DayNumb+1).font =  fontstyle
                                #SH.column_dimensions[SH.cell(row = r+1, column = DayNumb+1).column].width = 22
                                SH.cell(row = r+1, column = DayNumb+1).alignment = Alignment(horizontal = 'center', vertical = 'center')                        
                            if r == 1: 
                                SH.cell(row = r+1 , column = DayNumb+1).value = DaysList[(dt.weekday())]
                                SH.cell(row = r+1, column = DayNumb+1).font =  fontstyle                        
                                SH.cell(row = r+1, column = DayNumb+1).alignment = Alignment(horizontal = 'center', vertical = 'center')                                                
                                print(DaysList[(dt.weekday())])
                            if r >=2 and SH.cell(row =  2, column = DayNumb+1).value != 'Sunday':
                                SH.cell(row = r+1, column = DayNumb+1).value = 'A'
                                SH.cell(row = r+1, column = DayNumb+1).alignment = Alignment(horizontal = 'center', vertical = 'center')                        
                                
                SH.title = f"{dt:%B}" + "  " + str(Year)         
        except ValueError:
            messagebox.showinfo('warning',"minimum value must be smaller than maximum value")

        finally:
            wkb.save(Wkbpath)
            messagebox.showinfo('warning',"Creation successfull")


    btn=ttk.Button(window,text="submit",width=12,command=action)
    btn.grid(row=12,column=3)


def populate_list():
    attendance_list.delete(0,END)
    for row in db.fetch():
        attendance_list.insert(END,row)
        

def clear_list():
     attendance_list.delete(0,END)


#attendance list taken
attendance_list=Listbox(window,height=8,width=50,border=0)
attendance_list.grid(row=4,column=0,columnspan=3,rowspan=6,pady=20,padx=20)

#scroll bar
scrollbar=Scrollbar(window)
scrollbar.grid(row=4,column=3)

#set scroll to listbox
attendance_list.configure(yscrollcommand=scrollbar.set)
scrollbar.configure(command=attendance_list.yview)

recognize=ttk.Button(window,text="recognize faces",command=recognize)
recognize.grid(row=0,column=0)
train=ttk.Button(window,text="train faces",command=train)
train.grid(row=0,column=1)
attendance=ttk.Button(window,text="take attendance",command=take_attendance)
attendance.grid(row=0,column=3)

excel=ttk.Button(window,text="Generate excel sheet",command=generate_excel)
excel.grid(row=0,column=2)

clrlist =ttk.Button(window,text="clear list",command=clear_list)
clrlist.grid(row=2,column=2)


populate_list()
window.mainloop()